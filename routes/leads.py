from flask import Blueprint, render_template, request, send_file, redirect, url_for, jsonify, flash
import pandas as pd
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font
from pathlib import Path
import os
import csv
import io
from typing import List, Tuple
from datetime import datetime

# Import services with error handling
try:
    from models.database import db
except ImportError:
    db = None

try:
    from services.ollama_service import ollama_service
except ImportError:
    ollama_service = None

try:
    from services.serp_service import serp_service
except ImportError:
    serp_service = None

try:
    from config import EXPORT_FOLDER, SCIHUB_FOLDER, SERP_ENGINES, DEFAULT_RESEARCH_QUESTION
except ImportError:
    EXPORT_FOLDER = "exports"
    SCIHUB_FOLDER = "scihub"
    SERP_ENGINES = ["google"]
    DEFAULT_RESEARCH_QUESTION = "epigenetik och pre-diabetes"

try:
    from utils.performance import get_session
except ImportError:
    get_session = None

try:
    from utils.logger import get_logger
    logger = get_logger('leads')
except ImportError:
    logger = None

leads_bp = Blueprint('leads', __name__)

@leads_bp.route('/')
def show_leads():
    """Display all leads with enhanced table"""
    if not db:
        return "Database not available", 500
    
    # Get pagination parameters
    page = request.args.get('page', 1, type=int)
    per_page = 20
    offset = (page - 1) * per_page
    
    # Get all leads for statistics
    all_leads = db.get_all_leads()
    total_leads = len(all_leads)
    
    # Calculate pagination
    total_pages = (total_leads + per_page - 1) // per_page
    start_index = offset
    end_index = min(offset + per_page, total_leads)
    
    # Get paginated leads
    leads = all_leads[start_index:end_index]
    
    # Calculate statistics
    high_quality_count = sum(1 for lead in all_leads if lead.get('ai_summary'))
    filtered_count = len(leads)
    
    # Get search history for recent activity
    search_history = db.get_search_history(5) if db else []
    
    # Ensure ollama status is checked
    if ollama_service:
        ollama_status = ollama_service.check_status()
    else:
        ollama_status = {"ok": False, "msg": "Ollama service not available"}
    
    # Check AutoGPT availability
    try:
        from leadfinder_autogpt_integration import LeadfinderAutoGPTIntegration
        autogpt_integration = LeadfinderAutoGPTIntegration("mistral:latest")
        autogpt_available = True
    except Exception as e:
        autogpt_available = False
    
    return render_template('leads_enhanced.html', 
                         leads=leads, 
                         total_leads=total_leads,
                         filtered_count=filtered_count,
                         high_quality_count=high_quality_count,
                         search_history=search_history,
                         ollama_status=ollama_status,
                         engines=SERP_ENGINES,
                         selected_engines=["google"],
                         autogpt_available=autogpt_available,
                         # Pagination data
                         current_page=page,
                         total_pages=total_pages,
                         start_index=start_index,
                         end_index=end_index,
                         per_page=per_page)

@leads_bp.route('/export')
def export_to_excel():
    """Export leads to Excel file with enhanced information"""
    if not db:
        return "Database not available", 500
    
    leads = db.get_all_leads()
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Leads"
    headers = ["ID", "Title", "Description", "Link", "AI Summary", "Source", "Tags", 
              "Company", "Institution", "Contact Name", "Contact Email", "Contact Phone", 
              "Contact LinkedIn", "Contact Status", "Notes", "Created", "Updated"]
    ws.append(headers)
    
    for lead in leads:
        row = [
            lead.get('id'),
            lead.get('title'),
            lead.get('description'),
            lead.get('link'),
            lead.get('ai_summary'),
            lead.get('source'),
            lead.get('tags'),
            lead.get('company'),
            lead.get('institution'),
            lead.get('contact_name'),
            lead.get('contact_email'),
            lead.get('contact_phone'),
            lead.get('contact_linkedin'),
            lead.get('contact_status'),
            lead.get('notes'),
            lead.get('created_at'),
            lead.get('updated_at')
        ]
        ws.append(row)
        # Make link clickable
        if lead.get('link'):
            link_cell = ws.cell(row=ws.max_row, column=4)
            link_cell.hyperlink = lead.get('link')
            link_cell.font = Font(color="0000FF", underline="single")
    
    # Set column width
    for i, col in enumerate(headers, 1):
        ws.column_dimensions[get_column_letter(i)].width = 25
    
    filename = "leads_export_enhanced.xlsx"
    wb.save(filename)
    return send_file(filename, as_attachment=True)

@leads_bp.route('/export/csv')
def export_to_csv():
    """Export leads to CSV file with enhanced information"""
    if not db:
        return "Database not available", 500
    
    leads = db.get_all_leads()
    
    # Create CSV in memory
    output = io.StringIO()
    writer = csv.writer(output)
    
    # Write headers
    headers = ["ID", "Title", "Description", "Link", "AI Summary", "Source", "Tags", 
              "Company", "Institution", "Contact Name", "Contact Email", "Contact Phone", 
              "Contact LinkedIn", "Contact Status", "Notes", "Created", "Updated"]
    writer.writerow(headers)
    
    # Write data
    for lead in leads:
        row = [
            lead.get('id'),
            lead.get('title'),
            lead.get('description'),
            lead.get('link'),
            lead.get('ai_summary'),
            lead.get('source'),
            lead.get('tags'),
            lead.get('company'),
            lead.get('institution'),
            lead.get('contact_name'),
            lead.get('contact_email'),
            lead.get('contact_phone'),
            lead.get('contact_linkedin'),
            lead.get('contact_status'),
            lead.get('notes'),
            lead.get('created_at'),
            lead.get('updated_at')
        ]
        writer.writerow(row)
    
    # Prepare response
    output.seek(0)
    return send_file(
        io.BytesIO(output.getvalue().encode('utf-8')),
        mimetype='text/csv',
        as_attachment=True,
        download_name='leads_export_enhanced.csv'
    )

@leads_bp.route('/download_links')
def download_links():
    """Download content from all links in Excel file"""
    logs = []
    excel_path = 'leads_export.xlsx'
    
    try:
        os.makedirs(EXPORT_FOLDER, exist_ok=True)
        logs.append(f"Mapp OK: {EXPORT_FOLDER}")
    except Exception as e:
        logs.append(f"<span style='color:red'>Kunde inte skapa mapp: {EXPORT_FOLDER} - {e}</span>")
        return "<br>".join(logs)
    
    try:
        df = pd.read_excel(excel_path)
        logs.append(f"Excel OK: {excel_path}")
        logs.append(f"Kolumner: {list(df.columns)}")
        logs.append(f"Antal rader: {len(df)}")
    except Exception as e:
        logs.append(f"<span style='color:red'>Kunde inte läsa Excel: {e}</span>")
        return "<br>".join(logs)
    
    if 'Länk' not in df.columns:
        logs.append('<b>Ingen kolumn "Länk" hittades i Excel-filen.</b>')
        return "<br>".join(logs)
    
    # Use optimized session if available
    if get_session:
        session = get_session()
    else:
        import requests
        session = requests
    
    n_links = 0
    for _, row in df.iterrows():
        url = str(row['Länk'])
        if isinstance(url, str) and url.startswith("http"):
            n_links += 1
            
            # Sanitize filename to prevent path traversal and invalid characters
            try:
                # Extract filename from URL and sanitize it
                raw_filename = url.split("/")[-1].split("?")[0]
                if not raw_filename or raw_filename == "":
                    raw_filename = f"lead_{row['ID']}.html"
                
                # Remove any path separators and invalid characters
                import re
                # Remove path separators, control characters, and other dangerous chars
                safe_filename = re.sub(r'[<>:"/\\|?*\x00-\x1f]', '_', raw_filename)
                # Limit length to prevent issues
                safe_filename = safe_filename[:100]
                # Ensure it has a valid extension
                if not safe_filename.endswith(('.html', '.htm', '.pdf', '.txt')):
                    safe_filename += '.html'
                
                # Additional safety check - ensure filename is not empty or just dots
                if not safe_filename or safe_filename in ['.', '..']:
                    safe_filename = f"lead_{row['ID']}.html"
                
                filename = safe_filename
                
                r = session.get(url, timeout=10)
                r.raise_for_status()
                
                # Use Path to safely join and prevent path traversal
                file_path = Path(EXPORT_FOLDER) / filename
                # Additional safety check - ensure file is within export folder
                if not str(file_path.resolve()).startswith(str(Path(EXPORT_FOLDER).resolve())):
                    logs.append(f"<span style='color:red'>Säkerhetsfel:</span> Ogiltigt filnamn: {filename}")
                    continue
                
                with open(file_path, "wb") as f:
                    f.write(r.content)
                logs.append(f"<span style='color:green'>Nedladdad:</span> {filename}")
                
            except Exception as e:
                logs.append(f"<span style='color:red'>Misslyckades:</span> {url} - {e}")
    
    if n_links == 0:
        logs.append('<b>Inga giltiga länkar hittades i kolumnen "Länk".</b>')
    
    return "<br>".join(logs)

@leads_bp.route('/delete/<int:lead_id>', methods=['POST'])
def delete_lead(lead_id: int):
    """Delete a specific lead"""
    if not db:
        return "Database not available", 500
    
    if db.delete_lead(lead_id):
        return redirect(url_for('leads.show_leads'))
    else:
        return "Error deleting lead", 400

@leads_bp.route('/bulk-delete', methods=['POST'])
def bulk_delete_leads():
    """Delete multiple leads at once"""
    if not db:
        return "Database not available", 500
    
    lead_ids = request.form.getlist('lead_ids')
    if not lead_ids:
        return "No leads selected", 400
    
    try:
        # Convert to integers and delete each lead
        deleted_count = 0
        for lead_id_str in lead_ids:
            try:
                lead_id = int(lead_id_str)
                if db.delete_lead(lead_id):
                    deleted_count += 1
            except ValueError:
                continue
        
        if deleted_count > 0:
            flash(f'Successfully deleted {deleted_count} lead(s)', 'success')
        else:
            flash('No leads were deleted', 'warning')
            
    except Exception as e:
        if logger:
            logger.error(f"Error in bulk delete: {e}")
        flash(f'Error deleting leads: {e}', 'error')
    
    return redirect(url_for('leads.show_leads'))

@leads_bp.route('/leads_by_source/<source>')
def leads_by_source(source: str):
    """Show leads filtered by source"""
    if not db:
        leads = []
    else:
        leads = db.get_leads_by_source(source)
    
    # Ensure ollama status is checked
    if ollama_service:
        ollama_status = ollama_service.check_status()
    else:
        ollama_status = {"ok": False, "msg": "Ollama service not available"}
    
    return render_template('leads.html', 
                         leads=leads, 
                         lead_count=len(leads),
                         current_source=source,
                         ollama_status=ollama_status,
                         engines=SERP_ENGINES,
                         selected_engines=["google"])

@leads_bp.route('/summarize/<int:lead_id>', methods=['POST'])
def summarize_lead(lead_id: int):
    """Summarize a lead with AI and update its summary in the database"""
    if not db or not ollama_service:
        return "Database or AI service not available", 500

    # Fetch the lead
    leads = db.get_all_leads()
    lead = next((l for l in leads if l[0] == lead_id), None)
    if not lead:
        return "Lead not found", 404

    title = lead[1]
    snippet = lead[2]
    link = lead[3]
    research_question = lead[4] if lead[4] else DEFAULT_RESEARCH_QUESTION

    # Generate summary with AI
    try:
        ai_summary = ollama_service.analyze_relevance(title, snippet, link, research_question)
        if not ai_summary:
            ai_summary = "AI kunde inte generera en sammanfattning."
    except Exception as e:
        ai_summary = f"AI-fel: {e}"

    # Update the lead in the database
    try:
        with db._get_connection() as conn:
            c = conn.cursor()
            c.execute('UPDATE leads SET ai_summary = ? WHERE id = ?', (ai_summary, lead_id))
            conn.commit()
    except Exception as e:
        return f"Kunde inte uppdatera lead: {e}", 500

    return redirect(url_for('leads.show_leads'))

@leads_bp.route('/summarize_ajax/<int:lead_id>', methods=['POST'])
def summarize_lead_ajax(lead_id: int):
    """AJAX endpoint for lead summarization"""
    if not db:
        return jsonify({'error': 'Database not available'}), 500
    
    try:
        lead = db.get_lead(lead_id)
        if not lead:
            return jsonify({'error': 'Lead not found'}), 404
        
        # Get AI summary
        if ollama_service:
            summary = ollama_service.analyze_relevance(
                lead['title'], 
                lead['snippet'], 
                lead['link'], 
                "general relevance"
            )
            
            # Update lead with summary
            db.update_lead(lead_id, {'ai_summary': summary})
            
            return jsonify({
                'success': True,
                'summary': summary,
                'lead_id': lead_id
            })
        else:
            return jsonify({'error': 'AI service not available'}), 503
            
    except Exception as e:
        if logger:
            logger.error(f"Lead summarization failed: {e}")
        return jsonify({'error': f'Summarization failed: {str(e)}'}), 500


# REST API Endpoints
@leads_bp.route('/api/leads', methods=['GET'])
def get_leads_api():
    """REST API endpoint for getting all leads"""
    if not db:
        return jsonify({'error': 'Database not available'}), 500
    
    try:
        page = request.args.get('page', 1, type=int)
        per_page = request.args.get('per_page', 20, type=int)
        source = request.args.get('source')
        
        leads = db.get_all_leads()
        
        # Filter by source if specified
        if source:
            leads = [lead for lead in leads if lead.get('source') == source]
        
        # Pagination
        total = len(leads)
        start = (page - 1) * per_page
        end = start + per_page
        paginated_leads = leads[start:end]
        
        return jsonify({
            'leads': paginated_leads,
            'pagination': {
                'page': page,
                'per_page': per_page,
                'total': total,
                'pages': (total + per_page - 1) // per_page
            }
        })
        
    except Exception as e:
        if logger:
            logger.error(f"API leads fetch failed: {e}")
        return jsonify({'error': f'Failed to fetch leads: {str(e)}'}), 500


@leads_bp.route('/api/leads/<int:lead_id>', methods=['GET'])
def get_lead_api(lead_id: int):
    """REST API endpoint for getting a specific lead"""
    if not db:
        return jsonify({'error': 'Database not available'}), 500
    
    try:
        lead = db.get_lead(lead_id)
        if not lead:
            return jsonify({'error': 'Lead not found'}), 404
        
        return jsonify({'lead': lead})
        
    except Exception as e:
        if logger:
            logger.error(f"API lead fetch failed: {e}")
        return jsonify({'error': f'Failed to fetch lead: {str(e)}'}), 500


@leads_bp.route('/api/leads/<int:lead_id>', methods=['PUT'])
def update_lead_api(lead_id: int):
    """REST API endpoint for updating a lead"""
    if not db:
        return jsonify({'error': 'Database not available'}), 500
    
    try:
        data = request.get_json()
        if not data:
            return jsonify({'error': 'No data provided'}), 400
        
        success = db.update_lead(
            lead_id=lead_id,
            title=data.get('title'),
            description=data.get('description'),
            link=data.get('link'),
            ai_summary=data.get('ai_summary'),
            source=data.get('source'),
            tags=data.get('tags'),
            company=data.get('company'),
            institution=data.get('institution'),
            contact_name=data.get('contact_name'),
            contact_email=data.get('contact_email'),
            contact_phone=data.get('contact_phone'),
            contact_linkedin=data.get('contact_linkedin'),
            contact_status=data.get('contact_status'),
            notes=data.get('notes')
        )
        
        if success:
            return jsonify({
                'success': True,
                'message': 'Lead updated successfully',
                'lead_id': lead_id
            }), 200
        else:
            return jsonify({'error': 'Lead not found'}), 404
            
    except Exception as e:
        if logger:
            logger.error(f"API lead update failed: {e}")
        return jsonify({'error': f'Failed to update lead: {str(e)}'}), 500

@leads_bp.route('/api/leads/<int:lead_id>', methods=['DELETE'])
def delete_lead_api(lead_id: int):
    """REST API endpoint for deleting a lead"""
    if not db:
        return jsonify({'error': 'Database not available'}), 500
    
    try:
        lead = db.get_lead(lead_id)
        if not lead:
            return jsonify({'error': 'Lead not found'}), 404
        
        db.delete_lead(lead_id)
        
        return jsonify({
            'success': True,
            'message': 'Lead deleted successfully',
            'lead_id': lead_id
        })
        
    except Exception as e:
        if logger:
            logger.error(f"API lead deletion failed: {e}")
        return jsonify({'error': f'Failed to delete lead: {str(e)}'}), 500


@leads_bp.route('/api/leads/export', methods=['GET'])
def export_leads_api():
    """REST API endpoint for exporting leads as JSON"""
    if not db:
        return jsonify({'error': 'Database not available'}), 500
    
    try:
        leads = db.get_all_leads()
        
        return jsonify({
            'leads': leads,
            'exported_at': datetime.now().isoformat(),
            'total_leads': len(leads)
        })
        
    except Exception as e:
        if logger:
            logger.error(f"API leads export failed: {e}")
        return jsonify({'error': f'Failed to export leads: {str(e)}'}), 500 